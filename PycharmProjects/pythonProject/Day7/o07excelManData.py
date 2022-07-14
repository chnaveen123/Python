
from openpyxl import load_workbook

wb = load_workbook("1FirstExcel.xlsx")

wb.active = wb['Player']

ws = wb.active

for row in ws.iter_rows(min_row=2, min_col=1, max_col=7):
    for cell in row:
        if cell.value == "Lara":
            cell.value = "brain lara".upper()

wb.save("1FirstExcel.xlsx")


